"""Self-contained .xlsx builders for the download module.

Kept dependency-light and decoupled from the rest of the app on purpose: the
whole download feature (this module + routers/download.py) is meant to be lifted
into its own project later, so it only needs openpyxl plus the ORM rows passed in.

Returns raw bytes so the router can stream them with a Content-Disposition header.
"""
from __future__ import annotations

import io
from datetime import datetime
from decimal import Decimal
from typing import TYPE_CHECKING, Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:  # imports only for type hints — keeps this module ORM-agnostic
    from app.models import Record, ScreenSnapshot, ScreenSnapshotItem

_HEADER_FILL = PatternFill("solid", fgColor="FCE4D6")
_SESSION_LABEL = {"intraday_1300": "盤中13:00", "eod": "收盤後"}


def _num(value: Decimal | float | None) -> float | None:
    return None if value is None else float(value)


def _autosize(ws, widths: Sequence[int]) -> None:
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def _style_header(ws, row_idx: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        c = ws.cell(row=row_idx, column=col)
        c.font = Font(bold=True)
        c.fill = _HEADER_FILL
        c.alignment = Alignment(horizontal="center")


def _workbook_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Screening snapshot -> xlsx
# ---------------------------------------------------------------------------

_SNAPSHOT_HEADERS = [
    "排名", "代號", "名稱", "市場", "現價", "漲跌", "漲幅%", "量(張)",
    "昨高", "量比", "5MA", "月線(20MA)", "月線上彎", "昨收",
]
_SNAPSHOT_WIDTHS = [6, 8, 12, 6, 9, 8, 8, 9, 9, 7, 9, 11, 9, 9]


def snapshot_to_xlsx(
    snapshot: ScreenSnapshot, items: Iterable[ScreenSnapshotItem]
) -> bytes:
    items = list(items)
    wb = Workbook()
    ws = wb.active
    ws.title = "起漲篩選"

    label = _SESSION_LABEL.get(snapshot.session, snapshot.session)
    title = (
        f"{snapshot.trade_date:%Y-%m-%d} {label} 起漲篩選"
        f"（紅K+突破昨高+量增+站上5MA+月線上彎+昨日在5MA下） 共 {len(items)} 檔"
    )
    ws.append([title])
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.merge_cells(
        start_row=1, start_column=1, end_row=1, end_column=len(_SNAPSHOT_HEADERS)
    )

    ws.append(_SNAPSHOT_HEADERS)
    _style_header(ws, 2, len(_SNAPSHOT_HEADERS))

    for it in items:
        ws.append([
            it.rank,
            it.symbol,
            it.name,
            it.market,
            _num(it.close),
            _num(it.change),
            _num(it.change_pct),
            _num(it.lots),
            _num(it.prev_high),
            _num(it.vol_ratio),
            _num(it.ma5),
            _num(it.ma20),
            "↑" if it.ma20_up else "",
            _num(it.prev_close),
        ])

    first = 3
    last = first + len(items) - 1
    if items:
        for row in ws.iter_rows(min_row=first, max_row=last):
            for c in (row[4], row[5], row[6], row[8], row[9], row[10], row[11], row[13]):
                c.number_format = "0.00"
            row[12].alignment = Alignment(horizontal="center")

    _autosize(ws, _SNAPSHOT_WIDTHS)
    ws.freeze_panes = "A3"
    return _workbook_bytes(wb)


def empty_snapshot_xlsx(trade_date_label: str, session: str) -> bytes:
    """A valid, clearly-labelled workbook for a date+session with no results."""
    wb = Workbook()
    ws = wb.active
    ws.title = "起漲篩選"
    label = _SESSION_LABEL.get(session, session)
    ws.append([f"{trade_date_label} {label} 起漲篩選：今日無符合條件的個股"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    return _workbook_bytes(wb)


# ---------------------------------------------------------------------------
# User records -> xlsx
# ---------------------------------------------------------------------------

_RECORD_HEADERS = ["代號", "名稱", "市場", "市場代碼", "目標價", "成本價", "最新收盤", "更新時間"]
_RECORD_WIDTHS = [8, 14, 8, 10, 10, 10, 10, 22]


def records_to_xlsx(records: Iterable[Record], owner_label: str = "") -> bytes:
    records = list(records)
    wb = Workbook()
    ws = wb.active
    ws.title = "我的紀錄"

    title = "我的個股紀錄"
    if owner_label:
        title += f"（{owner_label}）"
    title += f"　共 {len(records)} 檔　匯出 {datetime.now():%Y-%m-%d %H:%M}"
    ws.append([title])
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.merge_cells(
        start_row=1, start_column=1, end_row=1, end_column=len(_RECORD_HEADERS)
    )

    ws.append(_RECORD_HEADERS)
    _style_header(ws, 2, len(_RECORD_HEADERS))

    for r in records:
        ws.append([
            r.symbol,
            r.name,
            r.market,
            r.market_code,
            _num(r.target_price),
            _num(r.cost_price),
            _num(r.last_close),
            r.updated_at.strftime("%Y-%m-%d %H:%M") if r.updated_at else "",
        ])

    first = 3
    last = first + len(records) - 1
    if records:
        for row in ws.iter_rows(min_row=first, max_row=last):
            for c in (row[4], row[5], row[6]):
                c.number_format = "0.00"

    _autosize(ws, _RECORD_WIDTHS)
    ws.freeze_panes = "A3"
    return _workbook_bytes(wb)
