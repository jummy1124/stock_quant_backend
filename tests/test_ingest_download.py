import io

import pytest
from openpyxl import load_workbook

from app.config import settings
from tests.conftest import auth_header, register

TOKEN = "test-ingest-token"


@pytest.fixture(autouse=True)
def _set_ingest_token():
    original = settings.INGEST_TOKEN
    settings.INGEST_TOKEN = TOKEN
    yield
    settings.INGEST_TOKEN = original


def _payload(session="intraday_1300", trade_date="2026-06-24", n=2):
    items = [
        {
            "rank": i + 1,
            "symbol": f"23{i:02d}",
            "name": f"測試股{i}",
            "market": "上市",
            "market_code": "TWSE",
            "close": 100.5 + i,
            "prev_close": 95.0 + i,
            "change": 5.5,
            "change_pct": 5.78,
            "volume": 1234000,
            "lots": 1234.0,
            "open": 96.0,
            "high": 101.0,
            "low": 95.5,
            "prev_high": 98.0,
            "vol_ratio": 1.8,
            "ma5": 97.0,
            "ma20": 90.0,
            "ma20_up": True,
        }
        for i in range(n)
    ]
    return {
        "trade_date": trade_date,
        "session": session,
        "generated_at": f"{trade_date}T13:00:00+08:00",
        "source": "live",
        "universe": 1900,
        "quotable": 1850,
        "pool_size": 12,
        "warning": None,
        "items": items,
    }


def _ingest(client, payload, token=TOKEN):
    headers = {"X-Ingest-Token": token} if token is not None else {}
    return client.post("/userapi/ingest/snapshot", json=payload, headers=headers)


# ---------- ingest auth ----------


def test_ingest_requires_token(client):
    resp = _ingest(client, _payload(), token=None)
    assert resp.status_code == 401, resp.text


def test_ingest_rejects_wrong_token(client):
    resp = _ingest(client, _payload(), token="nope")
    assert resp.status_code == 401, resp.text


def test_ingest_503_when_unconfigured(client):
    settings.INGEST_TOKEN = ""
    resp = _ingest(client, _payload(), token="anything")
    assert resp.status_code == 503, resp.text


def test_ingest_rejects_bad_session(client):
    resp = _ingest(client, _payload(session="weekly"))
    assert resp.status_code == 422, resp.text


# ---------- ingest + idempotency ----------


def test_ingest_creates_then_replaces(client):
    resp = _ingest(client, _payload(n=2))
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["item_count"] == 2
    assert body["replaced"] is False

    # Re-ingest same date+session with a different count -> overwrite.
    resp2 = _ingest(client, _payload(n=3))
    assert resp2.status_code == 200, resp2.text
    body2 = resp2.json()
    assert body2["item_count"] == 3
    assert body2["replaced"] is True

    # Listing shows a single snapshot for that date+session.
    listing = client.get("/downloadapi/snapshots").json()["snapshots"]
    intraday = [s for s in listing if s["session"] == "intraday_1300"]
    assert len(intraday) == 1
    assert intraday[0]["item_count"] == 3


# ---------- snapshot listing + download (public) ----------


def test_snapshots_listing_public(client):
    _ingest(client, _payload(session="intraday_1300"))
    _ingest(client, _payload(session="eod"))
    resp = client.get("/downloadapi/snapshots")  # no auth
    assert resp.status_code == 200, resp.text
    sessions = {s["session"] for s in resp.json()["snapshots"]}
    assert sessions == {"intraday_1300", "eod"}


def test_download_snapshot_xlsx_public(client):
    _ingest(client, _payload(n=2))
    resp = client.get(
        "/downloadapi/snapshot.xlsx",
        params={"date": "2026-06-24", "session": "intraday_1300"},
    )
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats"
    )
    assert "attachment" in resp.headers["content-disposition"]
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    # title row + header row + 2 data rows
    assert ws.max_row == 4
    assert ws.cell(row=3, column=2).value == "2300"  # first symbol


def test_download_missing_snapshot_returns_empty_workbook(client):
    resp = client.get(
        "/downloadapi/snapshot.xlsx",
        params={"date": "2020-01-01", "session": "eod"},
    )
    assert resp.status_code == 200, resp.text
    wb = load_workbook(io.BytesIO(resp.content))
    assert "無符合" in (wb.active.cell(row=1, column=1).value or "")


def test_download_snapshot_bad_session(client):
    resp = client.get(
        "/downloadapi/snapshot.xlsx",
        params={"date": "2026-06-24", "session": "bogus"},
    )
    assert resp.status_code == 422, resp.text


# ---------- records download (requires JWT) ----------


def test_records_download_requires_auth(client):
    resp = client.get("/downloadapi/records.xlsx")
    assert resp.status_code == 401, resp.text


def test_records_download_returns_xlsx(client):
    headers = auth_header(client, email="dl@example.com")
    client.put(
        "/userapi/records/TWSE/2330",
        json={"name": "台積電", "market": "上市", "target_price": 1200,
              "cost_price": 1000, "last_close": 1100},
        headers=headers,
    )
    resp = client.get("/downloadapi/records.xlsx", headers=headers)
    assert resp.status_code == 200, resp.text
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    assert ws.cell(row=3, column=1).value == "2330"
    assert ws.cell(row=3, column=2).value == "台積電"
